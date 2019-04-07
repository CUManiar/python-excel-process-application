# Author: Chirag Maniar
# copyright @2019 by Chirag Maniar
# Execl sheet discount calculation and chart building
# Importing excel required packages
import openpyxl as excel


def process_excel_sheet(filename):

    # Selecting workbook i.e. Excel sheet
    workbook = excel.load_workbook(filename)

    # Selecting the sheet, operations are required to perform upon
    sheet = workbook["Sheet1"]

    # Calculating and adding discounted price in new created column for selected
    # row
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        discounted_price = cell.value * 0.9  # Discount of total 10%
        discounted_price_cell = sheet.cell(row, 4)  # Selecting the column and row
        discounted_price_cell.value = f"${discounted_price}"  # Put value after discount

    # Crate a new sheet with added discount
    workbook.save("Discount_Added_File.xlsx")
    print("File processed successfully!")


# Give excel file needs to be processed

process_excel_sheet("transactions.xlsx")
